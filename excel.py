import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Quantidade de linhas que serão geradas
QTD_LINHAS = 200

# Nome do arquivo final
ARQUIVO_SAIDA = "dados_aleatorios_voos.xlsx"


companhias = {
    "0": "Azul",
    "1": "Gol",
    "2": "LATAM",
    "3": "Copa Airlines",
    "4": "TAP Air Portugal",
    "5": "Avianca",
    "6": "American Airlines",
    "7": "Air France",
    "8": "Emirates",
    "9": "Lufthansa",
}

primeiros_nomes = [
    "João", "Maria", "Pedro", "Ana", "Lucas", "Mariana", "Gabriel", "Juliana",
    "Rafael", "Camila", "Bruno", "Fernanda", "Gustavo", "Larissa", "Felipe",
    "Beatriz", "Thiago", "Amanda", "Rodrigo", "Isabela"
]

nomes_do_meio = [
    "Henrique", "Eduardo", "Cristina", "Aparecida", "Roberto", "Carolina",
    "Miguel", "Vitória", "Antônio", "Luiza", "Ribeiro", "Almeida",
    "Ferreira", "Castro", "Moreira"
]

sobrenomes = [
    "Silva", "Santos", "Oliveira", "Souza", "Pereira", "Costa", "Rodrigues",
    "Almeida", "Nascimento", "Lima", "Araújo", "Fernandes", "Carvalho",
    "Gomes", "Martins"
]


def gerar_id_10_digitos():
    """
    Gera um número inteiro entre 0 e 9999999999.
    No Excel, ele será formatado com 10 dígitos, preservando zeros à esquerda.
    """
    return random.randint(0, 9_999_999_999)


def gerar_nome_brasileiro():
    primeiro = random.choice(primeiros_nomes)
    meio = random.choice(nomes_do_meio)
    sobrenome = random.choice(sobrenomes)
    return f"{primeiro} {meio} {sobrenome}"


# Cria o arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Dados"

# Cabeçalhos
cabecalhos = ["ID", "Companhia", "Embarque", "Nome"]
ws.append(cabecalhos)

# Gera os dados
for _ in range(QTD_LINHAS):
    id_numero = gerar_id_10_digitos()

    # Transforma em texto com 10 dígitos para conseguir ler o primeiro, segundo e terceiro dígitos
    id_texto = f"{id_numero:010d}"

    primeiro_digito = id_texto[0]
    segundo_terceiro = id_texto[1:3]

    companhia = companhias[primeiro_digito]
    nome = gerar_nome_brasileiro()

    ws.append([
        id_numero,
        companhia,
        segundo_terceiro,
        nome
    ])

# Formatação do cabeçalho
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Formata a coluna ID para sempre exibir 10 dígitos
for row in range(2, QTD_LINHAS + 2):
    ws[f"A{row}"].number_format = "0000000000"
    ws[f"C{row}"].alignment = Alignment(horizontal="center")

# Ajuste de largura das colunas
larguras = {
    "A": 15,
    "B": 22,
    "C": 12,
    "D": 30,
}

for coluna, largura in larguras.items():
    ws.column_dimensions[coluna].width = largura

# Congela a primeira linha
ws.freeze_panes = "A2"

# Salva o arquivo
wb.save(ARQUIVO_SAIDA)

print(f"Arquivo gerado com sucesso: {ARQUIVO_SAIDA}")